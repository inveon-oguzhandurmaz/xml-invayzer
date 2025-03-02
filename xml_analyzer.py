import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import threading
import os
from urllib.parse import urlparse
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import concurrent.futures
import asyncio
import aiohttp
import openpyxl

class XMLAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("XML Analiz Uygulaması")
        self.root.geometry("1000x700")
        self.root.configure(padx=20, pady=20)
        
        # Analiz kontrol değişkenleri
        self.analysis_paused = False
        self.analysis_stopped = False
        self.current_analysis_thread = None
        
        # Süre takibi için değişkenler
        self.analysis_start_time = None
        self.analysis_paused_time = None
        self.analysis_total_time = 0  # Duraklatılıp devam edildiğinde toplam süreyi hesaplamak için
        self.timer_running = False
        self.timer_id = None
        
        # Kaynak dosya adını tutmak için değişken ekle
        self.source_file_name = None
        # İndirilen XML dosyasının yolunu tutmak için değişken
        self.downloaded_xml_path = None
        
        # Uygulama kapatıldığında temizleme işlemi yap
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
        
        # Ana container frame oluştur
        main_container = ttk.Frame(root)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        self.xml_data = None
        self.tags_hierarchy = {}  # Tag hiyerarşisini tutacak sözlük
        self.results = []
        self.namespaces = {}  # XML namespace'leri tutmak için
        
        # Stil tanımlamaları
        self.style = ttk.Style()
        self.style.configure("TButton", font=('Helvetica', 10, 'bold'))
        self.style.configure("TLabel", font=('Helvetica', 11))
        
        # URL Giriş Kısmı
        url_frame = ttk.LabelFrame(main_container, text="XML URL Giriş", padding=(10, 5))
        url_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(url_frame, text="XML URL:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.url_entry = ttk.Entry(url_frame, width=60)
        self.url_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        self.download_btn = ttk.Button(url_frame, text="XML İndir", command=self.download_xml)
        self.download_btn.grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        
        # Veya Dosya Seçme
        file_frame = ttk.LabelFrame(main_container, text="Veya Yerel XML Seç", padding=(10, 5))
        file_frame.pack(fill=tk.X, pady=10)
        
        self.file_path_var = tk.StringVar()
        self.file_path_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=60, state="readonly")
        self.file_path_entry.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        
        self.browse_btn = ttk.Button(file_frame, text="Dosya Seç", command=self.browse_file)
        self.browse_btn.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        # İlerleme çubuğu
        self.progress_frame = ttk.LabelFrame(main_container, text="İşlem Durumu", padding=(10, 5))
        self.progress_frame.pack(fill=tk.X, pady=10)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.progress_frame, variable=self.progress_var, length=700, mode="determinate")
        self.progress_bar.pack(fill=tk.X, padx=5, pady=5)
        
        # Analiz süresi için etiket
        self.time_frame = ttk.Frame(self.progress_frame)
        self.time_frame.pack(fill=tk.X, padx=5)
        
        self.time_label = ttk.Label(self.time_frame, text="Geçen Süre:")
        self.time_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.elapsed_time_var = tk.StringVar()
        self.elapsed_time_var.set("00:00:00")
        self.elapsed_time_label = ttk.Label(self.time_frame, textvariable=self.elapsed_time_var)
        self.elapsed_time_label.pack(side=tk.LEFT)
        
        self.status_var = tk.StringVar()
        self.status_var.set("Hazır")
        self.status_label = ttk.Label(self.progress_frame, textvariable=self.status_var)
        self.status_label.pack(anchor=tk.W, padx=5, pady=5)
        
        # Üç adet tag seçme kısmı
        tag_frame = ttk.LabelFrame(main_container, text="XML Tag Seçimi", padding=(10, 5))
        tag_frame.pack(fill=tk.X, pady=10)
        
        # Base tag seçimi
        ttk.Label(tag_frame, text="Base Tag:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.base_tag_combobox = ttk.Combobox(tag_frame, width=20, state="disabled")
        self.base_tag_combobox.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        self.base_tag_combobox.bind("<<ComboboxSelected>>", self.on_base_tag_selected)
        
        # Parent tag seçimi
        ttk.Label(tag_frame, text="Parent Tag:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        self.parent_tag_combobox = ttk.Combobox(tag_frame, width=20, state="disabled")
        self.parent_tag_combobox.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)
        self.parent_tag_combobox.bind("<<ComboboxSelected>>", self.on_parent_tag_selected)
        
        # Child tag seçimi
        ttk.Label(tag_frame, text="Child Tag:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.child_tag_combobox = ttk.Combobox(tag_frame, width=20, state="disabled")
        self.child_tag_combobox.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        
        # Butonlar için frame
        button_frame = ttk.Frame(tag_frame)
        button_frame.grid(row=1, column=2, columnspan=2, sticky=tk.E, padx=5, pady=5)
        
        self.analyze_btn = ttk.Button(button_frame, text="Analiz Et", command=self.analyze_xml, state="disabled")
        self.analyze_btn.pack(side=tk.LEFT, padx=5)
        
        # Kontrol butonları
        self.pause_btn = ttk.Button(button_frame, text="Duraklat", command=self.pause_analysis, state="disabled")
        self.pause_btn.pack(side=tk.LEFT, padx=5)
        
        self.resume_btn = ttk.Button(button_frame, text="Devam Et", command=self.resume_analysis, state="disabled")
        self.resume_btn.pack(side=tk.LEFT, padx=5)
        
        self.stop_btn = ttk.Button(button_frame, text="Durdur", command=self.stop_analysis, state="disabled")
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        
        self.reset_btn = ttk.Button(button_frame, text="Sıfırla", command=self.reset_analysis, state="disabled")
        self.reset_btn.pack(side=tk.LEFT, padx=5)
        
        self.export_btn = ttk.Button(button_frame, text="Excel'e Aktar", command=self.export_to_excel, state="disabled")
        self.export_btn.pack(side=tk.LEFT, padx=5)
        
        # Sonuçlar için tablo frame'i
        result_frame = ttk.LabelFrame(main_container, text="Analiz Sonuçları", padding=(10, 5))
        result_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Treeview için scrollbar
        scrollbar_y = ttk.Scrollbar(result_frame, orient=tk.VERTICAL)
        scrollbar_x = ttk.Scrollbar(result_frame, orient=tk.HORIZONTAL)
        
        # Tablo oluşturma
        self.result_tree = ttk.Treeview(
            result_frame, 
            columns=("url", "status", "response_time"), 
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            selectmode="extended"  # Çoklu seçime izin ver
        )
        
        # Scrollbar'ları konumlandırma
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.config(command=self.result_tree.yview)
        scrollbar_x.config(command=self.result_tree.xview)
        
        # Tablo başlıkları
        self.result_tree.heading("url", text="Request URL")
        self.result_tree.heading("status", text="Response Status Code")
        self.result_tree.heading("response_time", text="Response Time")
        
        # Sütun genişlikleri
        self.result_tree.column("url", width=500, stretch=True)
        self.result_tree.column("status", width=150, anchor=tk.CENTER)
        self.result_tree.column("response_time", width=150, anchor=tk.CENTER)
        
        self.result_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Hücre seçimi için özel bağlama
        self.result_tree.bind("<Button-1>", self.on_tree_click)
        self.result_tree.bind("<Control-c>", self.copy_selected_cell)

    def browse_file(self):
        """Yerel XML dosyası seçme işlevi"""
        file_path = filedialog.askopenfilename(filetypes=[("XML Dosyaları", "*.xml")])
        if file_path:
            # Eğer daha önce indirilen bir XML dosyası varsa ve bu dosya seçilen dosyadan farklıysa, sil
            if self.downloaded_xml_path and os.path.exists(self.downloaded_xml_path) and self.downloaded_xml_path != file_path:
                try:
                    os.remove(self.downloaded_xml_path)
                    print(f"Önceki indirilen XML dosyası silindi: {self.downloaded_xml_path}")
                except Exception as e:
                    print(f"Dosya silinirken hata oluştu: {str(e)}")
                
                # İndirilen dosya yolunu sıfırla
                self.downloaded_xml_path = None
            
            self.file_path_var.set(file_path)
            self.load_xml_from_file(file_path)
    
    def download_xml(self):
        """URL'den XML indirme işlevi"""
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showerror("Hata", "Lütfen geçerli bir URL girin")
            return
        
        # URL'nin geçerli olup olmadığını kontrol et
        try:
            result = urlparse(url)
            if not all([result.scheme, result.netloc]):
                messagebox.showerror("Hata", "Geçerli bir URL giriniz")
                return
        except:
            messagebox.showerror("Hata", "URL ayrıştırma hatası")
            return
        
        # İndirme işlemini ayrı bir thread'de başlat
        self.status_var.set("XML indiriliyor...")
        self.progress_var.set(0)
        self.download_btn.config(state="disabled")
        self.browse_btn.config(state="disabled")
        
        threading.Thread(target=self._download_xml_thread, args=(url,), daemon=True).start()
    
    def _download_xml_thread(self, url):
        """XML indirme işlemi için thread"""
        try:
            self.status_var.set("XML indiriliyor...")
            
            # İndirme işlemi - yönlendirmeleri takip et ve stream olarak indir
            response = requests.get(url, stream=True, allow_redirects=True)
            if response.status_code != 200:
                self.root.after(0, lambda: messagebox.showerror("Hata", f"XML indirilemedi. Durum Kodu: {response.status_code}"))
                self.status_var.set("XML indirme hatası")
                self.root.after(0, self._reset_ui)
                return
            
            # Son URL'den dosya adını çıkar (yönlendirmeden sonra)
            final_url = response.url
            self.source_file_name = os.path.splitext(os.path.basename(urlparse(final_url).path))[0]
            if not self.source_file_name:
                self.source_file_name = "downloaded_xml"
            
            # Dosyayı orijinal adıyla kaydet
            xml_filename = f"{self.source_file_name}.xml"
            xml_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), xml_filename)
            
            # Toplam dosya boyutunu al (eğer sunucu destekliyorsa)
            total_size = int(response.headers.get('content-length', 0))
            
            # Eğer content-length yoksa, progress bar'ı belirsiz moda al
            if total_size == 0:
                self.progress_bar.config(mode="indeterminate")
                self.progress_bar.start()
                
                # Dosyayı indir
                with open(xml_file_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                
                # Progress bar'ı durdur ve normal moda al
                self.progress_bar.stop()
                self.progress_bar.config(mode="determinate")
                self.progress_var.set(100)
            else:
                # Progress bar'ı belirli moda al
                self.progress_bar.config(mode="determinate")
                
                # Dosyayı indir ve progress bar'ı güncelle
                downloaded_size = 0
                with open(xml_file_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                            downloaded_size += len(chunk)
                            progress = (downloaded_size / total_size) * 100
                            self.progress_var.set(progress)
                            self.status_var.set(f"İndiriliyor: {downloaded_size/1024:.1f} KB / {total_size/1024:.1f} KB ({int(progress)}%)")
                            self.root.update_idletasks()
            
            # İndirilen dosyanın yolunu kaydet
            self.downloaded_xml_path = xml_file_path
            self.status_var.set("XML başarıyla indirildi.")
            
            self.root.after(0, lambda: self.load_xml_from_file(xml_file_path))
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Hata", f"XML indirme hatası: {str(e)}"))
            self.status_var.set("Hata: " + str(e))
            self.root.after(0, self._reset_ui)
    
    def _reset_ui(self):
        """UI bileşenlerini sıfırla"""
        self.download_btn.config(state="normal")
        self.browse_btn.config(state="normal")
    
    def load_xml_from_file(self, file_path):
        """XML dosyasını yükle ve tag hiyerarşisini çıkar"""
        try:
            # Kaynak dosya adını kaydet
            self.source_file_name = os.path.splitext(os.path.basename(file_path))[0]
            
            tree = ET.parse(file_path)
            self.xml_data = tree
            root = tree.getroot()
            
            # RSS elementini atla ve içindeki ilk elementi root olarak kabul et
            if root.tag.lower().endswith('rss'):
                if len(root) > 0:
                    root = root[0]  # İlk child'ı root olarak kullan
            
            # Namespace'leri kaydet
            self.namespaces = {}
            for key, value in root.items():
                if key.startswith("{") or key.startswith("xmlns:"):
                    ns_parts = key.split("}")
                    if len(ns_parts) > 1:
                        ns = ns_parts[0].strip("{")
                        prefix = key.split(":")[1] if ":" in key else ""
                        self.namespaces[prefix] = ns
            
            # Tag hiyerarşisini oluştur
            self.tags_hierarchy = {}
            self._build_tag_hierarchy(root)
            
            # Root tag'i base tag listesine ekle
            root_tag = self._get_tag_name(root.tag)
            self.base_tag_combobox['values'] = [root_tag]
            self.base_tag_combobox.current(0)
            self.base_tag_combobox.config(state="readonly")
            
            # İlk base tag için parent tag'leri yükle
            self.on_base_tag_selected(None)
            
            self.status_var.set(f"XML yüklendi.")
            self.progress_var.set(100)
            self.reset_btn.config(state="normal")
            
        except Exception as e:
            messagebox.showerror("Hata", f"XML yükleme hatası: {str(e)}")
            self.status_var.set("XML yükleme hatası: " + str(e))
        finally:
            self.download_btn.config(state="normal")
            self.browse_btn.config(state="normal")
    
    def _get_tag_name(self, tag):
        """Namespace ile tag ismini ayırma"""
        if "}" in tag:
            return tag.split("}")[1]
        return tag
    
    def _build_tag_hierarchy(self, element, path=[]):
        """XML ağacındaki tag hiyerarşisini oluştur"""
        tag = self._get_tag_name(element.tag)
        current_path = path + [tag]
        
        # Mevcut yolun son tag'i için alt tagları topla
        if len(current_path) > 0:
            parent_path = '.'.join(current_path[:-1]) if len(current_path) > 1 else ""
            if parent_path not in self.tags_hierarchy:
                self.tags_hierarchy[parent_path] = []
            
            if tag not in self.tags_hierarchy[parent_path]:
                self.tags_hierarchy[parent_path].append(tag)
        
        # Alt elementlere devam et
        for child in element:
            self._build_tag_hierarchy(child, current_path)
    
    def on_base_tag_selected(self, event):
        """Base tag seçildiğinde parent tag'leri göster"""
        selected_base = self.base_tag_combobox.get()
        
        if selected_base:
            # Seçili base tag'in alt tag'leri
            parent_tags = self.tags_hierarchy.get("", [])
            parent_tags = [tag for tag in parent_tags if tag != selected_base]
            
            # Base tag'in doğrudan altındaki tag'ler
            direct_children = self.tags_hierarchy.get(selected_base, [])
            parent_tags.extend(direct_children)
            
            # Tekrarlananları kaldır ve sırala
            parent_tags = sorted(list(set(parent_tags)))
            
            if parent_tags:
                self.parent_tag_combobox['values'] = parent_tags
                self.parent_tag_combobox.current(0)
                self.parent_tag_combobox.config(state="readonly")
                self.on_parent_tag_selected(None)
            else:
                self.parent_tag_combobox.set("")
                self.parent_tag_combobox.config(state="disabled")
                self.child_tag_combobox.set("")
                self.child_tag_combobox.config(state="disabled")
                self.analyze_btn.config(state="disabled")
        else:
            self.parent_tag_combobox.set("")
            self.parent_tag_combobox.config(state="disabled")
    
    def on_parent_tag_selected(self, event):
        """Parent tag seçildiğinde child tag'leri göster"""
        selected_base = self.base_tag_combobox.get()
        selected_parent = self.parent_tag_combobox.get()
        
        if selected_parent:
            # Parent tag'in altındaki tag'ler
            parent_path = selected_base if selected_parent in self.tags_hierarchy.get("", []) else selected_parent
            child_path = selected_parent if parent_path == selected_base else f"{selected_base}.{selected_parent}"
            
            child_tags = self.tags_hierarchy.get(parent_path, []) + self.tags_hierarchy.get(child_path, [])
            
            # Aynı tag'i kaldır ve sırala
            child_tags = [tag for tag in child_tags if tag != selected_parent]
            child_tags = sorted(list(set(child_tags)))
            
            if child_tags:
                self.child_tag_combobox['values'] = child_tags
                self.child_tag_combobox.current(0)
                self.child_tag_combobox.config(state="readonly")
                self.analyze_btn.config(state="normal")
            else:
                self.child_tag_combobox.set("")
                self.child_tag_combobox.config(state="disabled")
                # Child tag olmasa bile seçili tag'leri analiz edebilmek için analiz düğmesini aktif bırak
                self.analyze_btn.config(state="normal")
        else:
            self.child_tag_combobox.set("")
            self.child_tag_combobox.config(state="disabled")
            self.analyze_btn.config(state="disabled")
    
    def reset_analysis(self):
        """Analiz sonuçlarını ve seçimleri sıfırla"""
        # İndirilen XML dosyasını sil
        if self.downloaded_xml_path and os.path.exists(self.downloaded_xml_path):
            try:
                os.remove(self.downloaded_xml_path)
                print(f"İndirilen XML dosyası silindi: {self.downloaded_xml_path}")
                self.downloaded_xml_path = None
            except Exception as e:
                print(f"Dosya silinirken hata oluştu: {str(e)}")
        
        # Tabloyu temizle
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        
        # Sonuçları temizle
        self.results = []
        
        # XML verilerini temizle
        self.xml_data = None
        self.tags_hierarchy = {}
        
        # URL girişini temizle
        self.url_entry.delete(0, tk.END)
        
        # Dosya yolu alanını temizle
        self.file_path_var.set("")
        
        # Export butonunu devre dışı bırak
        self.export_btn.config(state="disabled")
        
        # Tag seçimlerini temizle ve devre dışı bırak
        self.base_tag_combobox.set("")
        self.base_tag_combobox.config(state="disabled")
        self.parent_tag_combobox.set("")
        self.parent_tag_combobox.config(state="disabled")
        self.child_tag_combobox.set("")
        self.child_tag_combobox.config(state="disabled")
        
        # Analiz butonunu devre dışı bırak
        self.analyze_btn.config(state="disabled")
        
        # Kontrol butonlarını devre dışı bırak
        self.pause_btn.config(state="disabled")
        self.resume_btn.config(state="disabled") 
        self.stop_btn.config(state="disabled")
        
        # XML İndir ve Dosya Seç butonlarını aktif hale getir
        self.download_btn.config(state="normal")
        self.browse_btn.config(state="normal")
        
        # Süre sayacını sıfırla
        self._reset_timer()
        
        self.status_var.set("Analiz sıfırlandı.")
        self.progress_var.set(0)
    
    def analyze_xml(self):
        """Seçilen tag'leri analiz et ve içindeki URL'lere istek at"""
        if not self.xml_data:
            messagebox.showerror("Hata", "Önce bir XML dosyası yükleyin")
            return
        
        base_tag = self.base_tag_combobox.get()
        parent_tag = self.parent_tag_combobox.get()
        child_tag = self.child_tag_combobox.get()
        
        if not base_tag:
            messagebox.showerror("Hata", "Lütfen en az bir base tag seçin")
            return
        
        # Analiz durumunu sıfırla
        self.analysis_paused = False
        self.analysis_stopped = False
        
        # UI'ı hazırla
        self.analyze_btn.config(state="disabled")
        self.base_tag_combobox.config(state="disabled")
        self.parent_tag_combobox.config(state="disabled")
        self.child_tag_combobox.config(state="disabled")
        self.download_btn.config(state="disabled")
        self.browse_btn.config(state="disabled")
        
        # Kontrol butonlarını aktifleştir
        self.pause_btn.config(state="normal")
        self.stop_btn.config(state="normal")
        self.resume_btn.config(state="disabled")
        
        self.progress_var.set(0)
        self.status_var.set("Analiz başlatılıyor...")
        
        # Süre sayacını sıfırla ve başlat
        self._reset_timer()
        self._start_timer()
        
        # Tabloyu temizle
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
            
        # Analiz işlemi için thread başlat
        self.current_analysis_thread = threading.Thread(
            target=self._analyze_thread, 
            args=(base_tag, parent_tag, child_tag), 
            daemon=True
        )
        self.current_analysis_thread.start()
    
    def _find_elements_by_path(self, base_tag, parent_tag=None, child_tag=None):
        """Belirtilen tag yoluna göre elementleri bul"""
        root = self.xml_data.getroot()
        root_tag = self._get_tag_name(root.tag)
        
        results = []
        
        # Base tag root ise doğrudan root'u kullan
        if base_tag == root_tag:
            base_elements = [root]
        else:
            # Tüm base tag'leri bul
            base_elements = []
            for elem in root.iter():
                if self._get_tag_name(elem.tag) == base_tag:
                    base_elements.append(elem)
        
        # Parent tag belirtilmemişse base elementleri dön
        if not parent_tag:
            return base_elements
        
        # Tüm base elementlerinin altındaki parent tag'leri bul
        parent_elements = []
        for base_elem in base_elements:
            for child in base_elem.iter():
                if self._get_tag_name(child.tag) == parent_tag:
                    parent_elements.append(child)
        
        # Child tag belirtilmemişse parent elementleri dön
        if not child_tag:
            return parent_elements
        
        # Tüm parent elementlerinin altındaki child tag'leri bul
        child_elements = []
        for parent_elem in parent_elements:
            for child in parent_elem.iter():
                if self._get_tag_name(child.tag) == child_tag:
                    child_elements.append(child)
        
        return child_elements
    
    def _analyze_thread(self, base_tag, parent_tag, child_tag):
        """XML analiz işlemi için thread"""
        try:
            # Browser benzeri headers
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'tr,en-US;q=0.7,en;q=0.3',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Sec-Fetch-User': '?1',
                'Cache-Control': 'max-age=0'
            }

            elements_to_analyze = []
            
            if child_tag:
                elements_to_analyze = self._find_elements_by_path(base_tag, parent_tag, child_tag)
            elif parent_tag:
                elements_to_analyze = self._find_elements_by_path(base_tag, parent_tag)
            else:
                elements_to_analyze = self._find_elements_by_path(base_tag)
            
            if not elements_to_analyze:
                tag_path = f"{base_tag}"
                if parent_tag:
                    tag_path += f" > {parent_tag}"
                if child_tag:
                    tag_path += f" > {child_tag}"
                self.root.after(0, lambda: messagebox.showinfo("Bilgi", f"'{tag_path}' yolunda element bulunamadı"))
                self.root.after(0, self._reset_analyze_ui)
                return
                
            total_elements = len(elements_to_analyze)
            self.status_var.set(f"{total_elements} adet element bulundu. URL'ler analiz ediliyor...")
            
            self.results = []
            
            # Önce tüm URL'leri topla
            all_urls = []
            for element in elements_to_analyze:
                # Element içinden URL'leri çıkar
                urls = self._extract_urls(element)
                
                # Eğer element metni doğrudan bir URL ise
                if not urls and element.text and (element.text.strip().startswith('http://') or element.text.strip().startswith('https://')):
                    urls = [element.text.strip()]
                
                all_urls.extend(urls)
            
            total_urls = len(all_urls)
            
            if total_urls == 0:
                self.root.after(0, lambda: messagebox.showinfo("Bilgi", "Seçilen elementlerde URL bulunamadı."))
                self.root.after(0, self._reset_analyze_ui)
                return
            
            self.status_var.set(f"Toplam {total_urls} URL bulundu. Analiz ediliyor...")
            
            # Asyncio event loop'unu başlat
            asyncio.run(self._analyze_urls_async(all_urls, headers))
            
            # Analiz tamamlandı
            if not self.analysis_stopped:
                self.status_var.set(f"Analiz tamamlandı. {len(self.results)} URL kontrol edildi.")
                self._pause_timer()  # Süre sayacını durdur
                self.export_btn.config(state="normal")
            elif self.analysis_paused:
                # Duraklatma durumu korunmalı
                pass
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Hata", f"Analiz hatası: {str(e)}"))
            self.status_var.set("Analiz hatası: " + str(e))
            self._pause_timer()  # Süre sayacını durdur
        finally:
            if not self.analysis_paused:
                self.root.after(0, self._reset_analyze_ui)
            self.current_analysis_thread = None
    
    async def _analyze_urls_async(self, urls, headers):
        """URL'leri asenkron olarak analiz et"""
        total_urls = len(urls)
        processed_urls = 0
        
        # Sonuçları saklamak için liste
        self.results = []
        
        # Semaphore ile eşzamanlı istek sayısını sınırla
        # Dengeli performans için optimum değer
        semaphore = asyncio.Semaphore(200)  # Aynı anda en fazla 200 istek
        
        # Yeniden deneme ayarları
        max_retries = 3
        retry_delay = 1  # saniye
        
        async def fetch_url(url, session):
            nonlocal processed_urls
            
            # Duraklatma ve durdurma kontrolü için fonksiyon
            async def check_pause_stop():
                # Duraklatma durumunda bekle
                while self.analysis_paused and not self.analysis_stopped:
                    # Duraklatma durumunda UI'ı güncelle (sadece ilk kez)
                    if not hasattr(check_pause_stop, 'paused_message_shown'):
                        self.root.after(0, lambda: self.status_var.set("Analiz duraklatıldı."))
                        check_pause_stop.paused_message_shown = True
                    await asyncio.sleep(0.1)
                
                # Durdurma durumunda
                if self.analysis_stopped:
                    # Durdurma durumunda UI'ı güncelle (sadece ilk kez)
                    if not hasattr(check_pause_stop, 'stopped_message_shown'):
                        self.root.after(0, lambda: self.status_var.set("Analiz durduruluyor..."))
                        check_pause_stop.stopped_message_shown = True
                    return True
                
                # Devam etme durumunda
                if hasattr(check_pause_stop, 'paused_message_shown'):
                    self.root.after(0, lambda: self.status_var.set("Analiz devam ediyor..."))
                    delattr(check_pause_stop, 'paused_message_shown')
                
                return False
            
            # Semaphore ile sınırlama
            async with semaphore:
                if await check_pause_stop():
                    return
                
                # Yeniden deneme mekanizması
                for retry in range(max_retries):
                    if await check_pause_stop():
                        return
                    
                    try:
                        start_time = time.time()
                        
                        # Önce HEAD isteği dene (daha hızlı)
                        try:
                            async with session.head(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30), allow_redirects=True) as response:
                                status_code = response.status
                                
                                # Eğer 403 alırsak GET ile deneyelim
                                if status_code == 403:
                                    async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30), allow_redirects=True) as response:
                                        status_code = response.status
                        except:
                            # HEAD isteği başarısız olursa GET ile dene
                            async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30), allow_redirects=True) as response:
                                status_code = response.status
                        
                        response_time = time.time() - start_time
                        
                        # Sonucu kaydet
                        result = (url, status_code, f"{response_time:.2f}s")
                        
                        # Thread-safe bir şekilde UI güncelle
                        self.root.after(0, lambda u=url, s=status_code, t=f"{response_time:.2f}s": 
                                       self.result_tree.insert("", tk.END, values=(u, s, t)))
                        
                        # Sonuçları listeye ekle
                        self.results.append(result)
                        
                        # İşlenen URL sayısını artır
                        processed_urls += 1
                        
                        # İlerleme çubuğunu güncelle
                        progress = (processed_urls / total_urls) * 100
                        self.root.after(0, lambda: self.progress_var.set(progress))
                        
                        # Sadece analiz durdurulmuyorsa durum mesajını güncelle
                        if not (self.analysis_stopped or self.analysis_paused):
                            self.root.after(0, lambda: self.status_var.set(
                                f"Analiz ediliyor: {processed_urls}/{total_urls} URL ({int(progress)}%)"
                            ))
                        break
                    
                    except asyncio.TimeoutError:
                        # Zaman aşımı hatası - yeniden dene
                        if retry < max_retries - 1:
                            await asyncio.sleep(retry_delay * (retry + 1))  # Artan bekleme süresi
                        else:
                            # Son deneme başarısız oldu
                            result = (url, "Hata: Zaman aşımı", "N/A")
                            self.root.after(0, lambda u=url, s="Hata: Zaman aşımı": 
                                           self.result_tree.insert("", tk.END, values=(u, s, "N/A")))
                            self.results.append(result)
                            processed_urls += 1
                            
                            # İlerleme çubuğunu güncelle
                            progress = (processed_urls / total_urls) * 100
                            self.root.after(0, lambda: self.progress_var.set(progress))
                            self.root.after(0, lambda: self.status_var.set(
                                f"Analiz ediliyor: {processed_urls}/{total_urls} URL ({int(progress)}%)"
                            ))
                    
                    except Exception as e:
                        # Diğer hatalar
                        error_msg = f"Hata: {str(e)}"
                        result = (url, error_msg, "N/A")
                        self.root.after(0, lambda u=url, s=error_msg: 
                                       self.result_tree.insert("", tk.END, values=(u, s, "N/A")))
                        self.results.append(result)
                        processed_urls += 1
                        
                        # İlerleme çubuğunu güncelle
                        progress = (processed_urls / total_urls) * 100
                        self.root.after(0, lambda: self.progress_var.set(progress))
                        self.root.after(0, lambda: self.status_var.set(
                            f"Analiz ediliyor: {processed_urls}/{total_urls} URL ({int(progress)}%)"
                        ))
                        break
        
        # Batch işleme için URL'leri gruplara böl
        # Dengeli performans için optimum değer
        batch_size = 2000  # Her seferde 2000 URL'yi işleme kuyruğuna al
        
        # Asenkron HTTP istemcisi oluştur
        # TCPConnector ile bağlantı havuzu oluştur ve DNS çözümlemesini önbelleğe al
        # limit=0: Bağlantı sayısı sınırı yok (semaphore ile kontrol ediliyor)
        # Daha uzun DNS önbellek süresi
        connector = aiohttp.TCPConnector(limit=0, ttl_dns_cache=1800)
        
        async with aiohttp.ClientSession(connector=connector) as session:
            for i in range(0, len(urls), batch_size):
                if self.analysis_stopped:
                    break
                
                batch_urls = urls[i:i+batch_size]
                
                # Her URL için bir görev oluştur
                tasks = []
                for url in batch_urls:
                    if self.analysis_stopped:
                        break
                    task = asyncio.create_task(fetch_url(url, session))
                    tasks.append(task)
                
                # Tüm görevlerin tamamlanmasını bekle
                if tasks:
                    await asyncio.gather(*tasks, return_exceptions=True)
                
                # Kısa bir bekleme ile sistem kaynaklarını rahatlatma
                await asyncio.sleep(0.1)
        
    def _extract_urls(self, element):
        """Bir element içinden URL'leri çıkarma"""
        urls = []
        
        # Element metnini kontrol et
        if element.text and ('http://' in element.text or 'https://' in element.text):
            # Basit URL çıkarma - gerçek durumda daha karmaşık olabilir
            text = element.text.strip()
            if text.startswith('http://') or text.startswith('https://'):
                urls.append(text)
        
        # Alt elementleri kontrol et
        for child in element:
            # Özyinelemeli olarak alt elementlerdeki URL'leri de kontrol et
            urls.extend(self._extract_urls(child))
            
        # Nitelikler (attributes) içindeki URL'leri kontrol et
        for attr_name, attr_value in element.attrib.items():
            if attr_value and ('http://' in attr_value or 'https://' in attr_value):
                if attr_value.startswith('http://') or attr_value.startswith('https://'):
                    urls.append(attr_value)
        
        return urls
    
    def _reset_analyze_ui(self):
        """Analiz UI'ını sıfırla"""
        self.analyze_btn.config(state="normal")
        self.base_tag_combobox.config(state="readonly")
        self.parent_tag_combobox.config(state="readonly")
        self.child_tag_combobox.config(state="readonly")
        self.download_btn.config(state="normal")
        self.browse_btn.config(state="normal")
        
        # Kontrol butonlarını devre dışı bırak
        self.pause_btn.config(state="disabled")
        self.resume_btn.config(state="disabled")
        self.stop_btn.config(state="disabled")
    
    def on_tree_click(self, event):
        """Treeview'da tıklama olayını yakala"""
        region = self.result_tree.identify("region", event.x, event.y)
        if region == "cell":
            column = self.result_tree.identify_column(event.x)
            column_index = int(column.replace("#", "")) - 1
            item = self.result_tree.identify_row(event.y)
            if item and column_index >= 0:
                # Tek bir hücreyi seç
                self.result_tree.selection_set(item)
                # Tıklanan sütunu hatırla
                self.selected_column = column_index
    
    def copy_selected_cell(self, event):
        """Seçili hücrenin içeriğini kopyala"""
        selection = self.result_tree.selection()
        if selection and hasattr(self, 'selected_column'):
            item = selection[0]
            values = self.result_tree.item(item, 'values')
            if values and len(values) > self.selected_column:
                # Seçili hücrenin değerini kopyala
                self.root.clipboard_clear()
                self.root.clipboard_append(str(values[self.selected_column]))
                self.status_var.set("Değer panoya kopyalandı")
                return "break"  # Varsayılan davranışı engelle
    
    def export_to_excel(self):
        """Sonuçları Excel olarak dışa aktar"""
        if not self.results:
            messagebox.showinfo("Bilgi", "Dışa aktarılacak sonuç yok")
            return
        
        # Varsayılan dosya adını oluştur
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"{self.source_file_name or 'xml'}_analyzed_{timestamp}.xlsx"
        
        # Dosya kaydetme iletişim kutusu
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx"), ("Tüm Dosyalar", "*.*")],
            initialfile=default_filename
        )
        
        if not file_path:
            return
        
        try:
            # Sonuçları DataFrame'e dönüştür
            df = pd.DataFrame(self.results, columns=["Request Url", "Response Status Code", "Response Time"])
            
            # Excel dosyasını oluştur
            workbook = Workbook()
            worksheet = workbook.active
            
            # Analiz süresini başa ekle
            if self.timer_running or self.elapsed_time_var.get() != "00:00:00":
                time_spent = self.elapsed_time_var.get()
            else:
                time_spent = "Belirsiz"
                
            worksheet.cell(row=1, column=1).value = f"Analiz Süresi: {time_spent}"
            worksheet.cell(row=1, column=1).font = Font(bold=True)
            worksheet.merge_cells('A1:C1')
            
            # Başlıkları ekle
            headers = ["Request Url", "Response Status Code", "Response Time"]
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=2, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Verileri ekle ve durum kodlarına göre renklendirme yap
            for row_num, result in enumerate(self.results, 3):
                # Durum koduna göre renklendirme için renk belirle
                status_code = result[1]
                if status_code == 200:
                    # Pastel yeşil
                    row_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif status_code == 404:
                    # Pastel turuncu
                    row_color = PatternFill(start_color="FFD8B0", end_color="FFD8B0", fill_type="solid")
                elif status_code == 500:
                    # Pastel kırmızı
                    row_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    # Pastel sarı
                    row_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                # URL - satırın ilk hücresi
                url_cell = worksheet.cell(row=row_num, column=1)
                url_cell.value = result[0]
                url_cell.fill = row_color
                
                # Durum kodu - satırın ikinci hücresi
                status_cell = worksheet.cell(row=row_num, column=2)
                status_cell.value = result[1]
                status_cell.alignment = Alignment(horizontal='center')
                status_cell.fill = row_color
                
                # Yanıt süresi - satırın üçüncü hücresi
                time_cell = worksheet.cell(row=row_num, column=3)
                time_cell.value = result[2]
                time_cell.alignment = Alignment(horizontal='center')
                time_cell.fill = row_color
            
            # Sütun genişliklerini otomatik ayarla
            for col in worksheet.columns:
                max_length = 0
                column = None
                
                # İlk hücreyi al ve sütun harfini belirle
                for cell in col:
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        column = cell.column_letter
                        break
                
                if column is None:
                    continue  # Sütunda birleştirilmemiş hücre yoksa atla
                
                # Hücre değerlerini kontrol et ve maksimum uzunluğu bul
                for cell in col:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # Maksimum genişliği ayarla (biraz ekstra boşluk bırak)
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width
            
            # Excel dosyasını kaydet
            workbook.save(file_path)
            
            self.status_var.set(f"Sonuçlar başarıyla dışa aktarıldı: {file_path}")
            
            # Excel dosyasını açma seçeneği ile popup göster
            result = messagebox.askquestion("Başarılı", 
                                           f"Sonuçlar başarıyla dışa aktarıldı:\n{file_path}\n\nExcel dosyasını açmak ister misiniz?",
                                           icon='info')
            if result == 'yes':
                # Excel dosyasını aç
                try:
                    os.startfile(file_path)
                except Exception as e:
                    messagebox.showerror("Hata", f"Excel dosyası açılırken hata oluştu: {str(e)}")
            
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dışa aktarma hatası: {str(e)}")
            self.status_var.set("Excel dışa aktarma hatası: " + str(e))

    def pause_analysis(self):
        """Analizi duraklat"""
        self.analysis_paused = True
        self.pause_btn.config(state="disabled")
        self.resume_btn.config(state="normal")
        self.status_var.set("Analiz duraklatılıyor...")
        self.export_btn.config(state="normal")
        
        # Süre sayacını duraklat
        self._pause_timer()
        
        # Durumun güncellenmesi için kısa bir bekleme
        self.root.after(1000, lambda: self.status_var.set("Analiz duraklatıldı."))

    def resume_analysis(self):
        """Analizi devam ettir"""
        self.analysis_paused = False
        self.pause_btn.config(state="normal")
        self.resume_btn.config(state="disabled")
        self.status_var.set("Analiz devam ediyor...")
        self.download_btn.config(state="disabled")
        
        # Süre sayacını devam ettir
        self._resume_timer()

    def stop_analysis(self):
        """Analizi durdur"""
        # Durdurma bayrağını ayarla
        self.analysis_stopped = True
        self.analysis_paused = False
        
        # Butonları güncelle
        self.pause_btn.config(state="disabled")
        self.resume_btn.config(state="disabled")
        self.stop_btn.config(state="disabled")
        
        # Durduruluyor mesajını göster
        self.status_var.set("Analiz durduruluyor...")
        
        # Süre sayacını duraklat
        self._pause_timer()
        
        # Analiz thread'i bitmesini bekle
        if self.current_analysis_thread and self.current_analysis_thread.is_alive():
            # Thread'in bitmesini bekle
            self.current_analysis_thread.join(timeout=0.5)  # En fazla 0.5 saniye bekle
            
            # UI güncellemesi için kısa bekleme
            self.root.after(2000, lambda: self.status_var.set("Analiz durduruldu."))
            self.root.after(2000, self._reset_analyze_ui)
        else:
            self.status_var.set("Analiz durduruldu.")
            self._reset_analyze_ui()

        self.export_btn.config(state="normal")

    def _on_closing(self):
        """Uygulama kapatıldığında çağrılır"""
        # İndirilen XML dosyasını sil
        if self.downloaded_xml_path and os.path.exists(self.downloaded_xml_path):
            try:
                os.remove(self.downloaded_xml_path)
                print(f"İndirilen XML dosyası silindi: {self.downloaded_xml_path}")
            except Exception as e:
                print(f"Dosya silinirken hata oluştu: {str(e)}")
        
        # Durdurma işlemini başlat
        self.analysis_stopped = True
        self.root.destroy()
    
    def _update_timer(self):
        """Süre sayacını günceller"""
        if not self.timer_running:
            return
        
        if self.analysis_paused:
            # Duraklatıldığında süreyi dondur
            elapsed_time = self.analysis_paused_time - self.analysis_start_time + self.analysis_total_time
        else:
            # Çalışırken süreyi güncelle
            current_time = time.time()
            elapsed_time = current_time - self.analysis_start_time + self.analysis_total_time
        
        # Süreyi formatlayarak göster
        hours, remainder = divmod(int(elapsed_time), 3600)
        minutes, seconds = divmod(remainder, 60)
        self.elapsed_time_var.set(f"{hours:02d}:{minutes:02d}:{seconds:02d}")
        
        # Her saniye güncelle
        self.timer_id = self.root.after(1000, self._update_timer)
    
    def _start_timer(self):
        """Süre sayacını başlatır"""
        if self.timer_running:
            return
            
        self.analysis_start_time = time.time()
        self.timer_running = True
        self._update_timer()
    
    def _pause_timer(self):
        """Süre sayacını duraklatır"""
        if not self.timer_running:
            return
            
        self.analysis_paused_time = time.time()
        self.timer_running = False
        
        # Timer'ı durdur
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
            self.timer_id = None
    
    def _resume_timer(self):
        """Süre sayacını devam ettirir"""
        if self.timer_running:
            return
            
        # Duraklatılmış süreyi toplam süreye ekle
        if self.analysis_paused_time:
            self.analysis_total_time += self.analysis_paused_time - self.analysis_start_time
        
        # Yeni başlangıç zamanı ayarla
        self.analysis_start_time = time.time()
        self.timer_running = True
        self._update_timer()
    
    def _reset_timer(self):
        """Süre sayacını sıfırlar"""
        self._pause_timer()
        self.analysis_start_time = None
        self.analysis_paused_time = None
        self.analysis_total_time = 0
        self.elapsed_time_var.set("00:00:00")


if __name__ == "__main__":
    root = tk.Tk()
    app = XMLAnalyzerApp(root)
    root.mainloop()